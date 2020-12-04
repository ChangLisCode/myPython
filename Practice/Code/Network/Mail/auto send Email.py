'''
Python Application 
Auto-send E-mail
> * get mail list for excel document
> * send prepared mail to all the mail address
> * mail with attachment, all document type
'''
# Source Code

import openpyxl, smtplib
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage

def send_mail(from_mail, mail_pass, recv):

    # 构建正文
    # 邮件正文,本内容用的是'plain',也可以使用'html'格式
    msg = MIMEText('''Dear Madame/Monsieur<br>I apply your position,please check my CV as attachment.<br>Best Wish<br>Yours
    ''',"html", "utf-8")  

    # 换行问题：如果邮件格式是文本格式plain,使用“\r\n”。如果件是HTML格式，使用"<br>"

    # 构建邮件附件
    # 添加附件（word文档）
    wordFile = '/Users/chang/Desktop/myCV.docx' # 附件文档路径
    word = MIMEApplication(open(wordFile, 'rb').read())
    word.add_header('Content-Disposition', 'attachment', filename='myCV.docx') #设置附件信息

    #添加附件（pdf文档）
    pdfFile = '/Users/chang/Desktop/Python_book.pdf'  #需发文件路径
    pdf = MIMEApplication(open(pdfFile, 'rb').read())
    pdf.add_header('Content-Disposition', 'attachment', filename='python_book.pdf')

    #添加图片
    imageFile = '/Users/chang/Desktop/图库/zenofpython.png'
    image = MIMEImage(open(imageFile, 'rb').read(), _subtype='octet-stream')
    image.add_header('Content-Disposition', 'attachment', filename='zen_of_python.png')
    
    m = MIMEMultipart()
    m.attach(msg)  #添加邮件正文内容
    m.attach(word)    #添加附件到邮件信息中
    m.attach(pdf)
    m.attach(image)
    m['Subject'] = 'supervisor position application' #邮件主题
    m['From'] = from_mail   #发件人
    m['To'] = recv    #收件人

    try:
        server = smtplib.SMTP('smtp.126.com')
        # 登陆邮箱（参数1：发件人邮箱，参数2：邮箱授权码）
        server.login(from_mail, mail_pass)
        # 发送邮件（参数1：发件人邮箱，参数2：若干收件人邮箱，参数3：把邮件内容格式改为str）
        server.sendmail(from_mail, recv.split(','), m.as_string())
        print('success')
        server.quit()
    except smtplib.SMTPException as e:
        print('error:', e)  # 打印错误


# 打开方式一 直接取值发送
wb = openpyxl.load_workbook('/Users/chang/Desktop/test1.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# 列表应用，取出 mail address
# mail_list = []

from_mail = input('input your account: ')       # 发件人邮箱
mail_pass = input('input your pwd.: ')          # 发件人邮箱验证码，到邮箱页面操作，非密码
# recv = mailAddress


for i in range(2, 5):
    recv = sheet.cell(row=i, column=2).value
    print(recv)

    # 带附件发送
    send_mail(from_mail, mail_pass, recv)
    print('mail send to '+ str(recv) + ' with attachment' + ' Done')


# 表格打开方式二 带筛选条件发送
wb = load_workbook('/Users/chang/Desktop/test1.xlsx')       # 写入邮件存储文件地址
sheet2 = wb['Sheet1']                                       # 文档表单

# 字典的应用，同时存入name和mail
mailList2 = {} 
for r in range(2, 6):
    payment = sheet.cell(row=r, column=4).value
    if payment != 1:
        name = sheet.cell(row=r, column=1).value # 获得name作为键
        email = sheet.cell(row=r, column=2).value # 获得email作为键值
        mailList2[name] = email # 添加到字典
print(mailList2)
